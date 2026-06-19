# Apple Serial Checker - v4.4 (moi serial = 1 trinh duyet moi, trang sach 100%)
import os, time, json, re, random, openpyxl
from collections import namedtuple
from openpyxl.styles import PatternFill, Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service 
from webdriver_manager.chrome import ChromeDriverManager

try:
    import undetected_chromedriver as uc
    HAS_UC = True
except Exception:
    HAS_UC = False
    print("!! Chưa cài undetected-chromedriver. Nên cài: pip install undetected-chromedriver")

RESULT_FILE = "results.xlsx"; INPUT_FILE = "serials.xlsx"

# ===== CẤU HÌNH PROXY =====
PROXY_KEY = ""        
PROXY_NHAMANG = "Random"  
PROXY_TINH = "0"          
PROXY_API_URL = ""        
PROXY_MIN_SEC = 62        
PROXY = ""                

try:
    if os.path.exists("proxy_key.txt"):
        _k = open("proxy_key.txt", encoding="utf-8").read().strip()
        if _k: PROXY_KEY = _k
except: pass

_my_ip = [None]
def my_public_ip():
    if _my_ip[0] is not None: return _my_ip[0]
    try:
        import requests
        _my_ip[0] = requests.get("https://api.ipify.org", timeout=10).text.strip()
    except: _my_ip[0] = ""
    return _my_ip[0]

def _build_proxy_url():
    if PROXY_API_URL: return PROXY_API_URL
    if PROXY_KEY:
        u = f"https://proxyxoay.shop/api/get.php?key={PROXY_KEY}&nhamang={PROXY_NHAMANG}&tinhthanh={PROXY_TINH}"
        wl = my_public_ip()
        if wl: u += f"&whitelist={wl}"
        return u
    return ""

BATCH_PER_IP = 1          # 1 IP / 1 serial (chống ban). KHÔNG load lại trang, chỉ load mã captcha mới khi chưa chắc
DELAY_MIN, DELAY_MAX = 2, 4

CAPTCHA_TRIES = 5 # Số lần LOAD MÃ captcha (refresh) để tìm mã đọc CHẮC. Hết mà chưa chắc -> đổi IP. KHÔNG load lại trang

def set_anti_sleep(status=True):
    try:
        import ctypes
        if status:
            ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001 | 0x00000002)
            print("   [+] Đã kích hoạt chế độ CHỐNG NGỦ.")
        else:
            ctypes.windll.kernel32.SetThreadExecutionState(0x80000000)
            print("   [-] Đã tắt chế độ chống ngủ.")
    except: pass

_last_change = [0.0]
def fetch_rotating_proxy():
    url = _build_proxy_url()
    if not url: return None
    wait = PROXY_MIN_SEC - (time.time() - _last_change[0])
    if wait > 0:
        print(f"   (Chờ {int(wait)}s cho đủ giới hạn đổi IP)...")
        time.sleep(wait + 0.5)
    try:
        import requests
        j = requests.get(url, timeout=45).json()
        if str(j.get("status")) != "100":
            print("   !! API proxy báo lỗi:", j.get("message") or j); return None
        raw = (j.get("proxyhttp") or "").strip()
        m = re.match(r"([\w.\-]+):(\d+)", raw)
        ph = f"{m.group(1)}:{m.group(2)}" if m else raw.replace(":", "", 0)
        _last_change[0] = time.time()
        print(f"   >> Proxy mới: {ph}  ({j.get('Nha Mang','')}/{j.get('Vi Tri','')})")
        return ph
    except Exception as e:
        print("   !! Lỗi gọi API proxy:", e); return None

UAS = [
 "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]
Result = namedtuple("Result", ["serial", "status", "model", "activation", "expiry"])

def read_serials(fp):
    wb = openpyxl.load_workbook(fp); ws = wb.active; out = []
    for row in ws.iter_rows(min_row=2, max_col=1):
        v = row[0].value
        if isinstance(v, str) and v.strip(): out.append(v.strip())
    return out

def load_existing_results(fp=RESULT_FILE):
    out = {}
    if not os.path.exists(fp): return out
    try:
        wb = openpyxl.load_workbook(fp); ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if not row or row[0].value is None: continue
            s = str(row[0].value).strip()
            if s:
                out[s] = Result(
                    serial=s, status=str(row[1].value or "").strip(),
                    model=str(row[2].value or ""), activation=str(row[3].value or ""),
                    expiry=str(row[4].value or "")
                )
    except: pass
    return out

def save_results(master_serials, existing_results, fp=RESULT_FILE):
    wb = openpyxl.Workbook(); ws = wb.active
    hdr = ['Serial', 'Status', 'Model', 'Ngay kich hoat', 'Ngay het BH']
    ws.append(hdr)
    for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="374151")
    
    colors = {"Activated": "C6EFCE", "Unactivated": "FFC7CE", "Invalid": "FFD1A4", "ERROR": "D9D9D9", "Unknown": "FFEB9C"}
    for s in master_serials:
        if s in existing_results:
            r = existing_results[s]
            ws.append([r.serial, r.status, r.model, r.activation, r.expiry])
            fg = colors.get(r.status, "FFFFFF")
            for c in ws[ws.max_row]: c.fill = PatternFill("solid", fgColor=fg)
        else:
            ws.append([s, "Chưa check", "", "", ""])
    for col, w in zip("ABCDE", [22, 13, 22, 16, 16]): ws.column_dimensions[col].width = w
    try: wb.save(fp)
    except PermissionError: print("   !! [!] CẢNH BÁO: Hãy ĐÓNG file Excel kết quả!")

def _parse_proxy(p):
    p = p.strip().replace("http://", "").replace("https://", "")
    user = pwd = ""
    if "@" in p:
        cred, hp = p.rsplit("@", 1)
        if ":" in cred: user, pwd = cred.split(":", 1)
    else: hp = p
    host, port = hp.split(":", 1) if ":" in hp else (hp, "80")
    return host, port, user, pwd

def _proxy_auth_ext(host, port, user, pwd, fn="proxy_auth_ext.zip"):
    import zipfile
    manifest = '{"version":"1.0","manifest_version":2,"name":"px","permissions":["proxy","tabs","unlimitedStorage","storage","<all_urls>","webRequest","webRequestBlocking"],"background":{"scripts":["bg.js"]},"minimum_chrome_version":"22.0.0"}'
    bg = ('var c={mode:"fixed_servers",rules:{singleProxy:{scheme:"http",host:"%s",port:parseInt(%s)},bypassList:["localhost"]}};'
        'chrome.proxy.settings.set({value:c,scope:"regular"},function(){});'
        'chrome.webRequest.onAuthRequired.addListener(function(d){return{authCredentials:{username:"%s",password:"%s"}};},{urls:["<all_urls>"]},["blocking"]);'
       ) % (host, port, user, pwd)
    z = zipfile.ZipFile(fn, "w"); z.writestr("manifest.json", manifest); z.writestr("bg.js", bg); z.close()
    return fn

def create_browser(proxy=None):
    ua = random.choice(UAS)
    host = port = user = pwd = ""
    if proxy: host, port, user, pwd = _parse_proxy(proxy)
    Opt = uc.ChromeOptions() if HAS_UC else webdriver.ChromeOptions()
    Opt.add_argument("--start-maximized")
    Opt.add_argument("--user-agent=" + ua)
    Opt.add_argument("--lang=vi-VN")
    if proxy and not user: Opt.add_argument("--proxy-server=http://%s:%s" % (host, port))
    if proxy and user:
        try: Opt.add_extension(_proxy_auth_ext(host, port, user, pwd))
        except: pass
    if HAS_UC:
        d = uc.Chrome(options=Opt, version_main=149)
    else:
        Opt.add_argument("--disable-blink-features=AutomationControlled")
        Opt.add_experimental_option("excludeSwitches", ["enable-automation"])
        Opt.add_experimental_option("useAutomationExtension", False)
        d = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=Opt)
    try: d.set_page_load_timeout(35)
    except: pass
    return d

def human_pause(driver):
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

def find_serial_input(driver, timeout=10):
    strategies = [
        (By.CSS_SELECTOR, ".serial-form-section #serial-number-input"),
        (By.CSS_SELECTOR, ".serial-form-section input.form-textbox-input"),
        (By.ID, "serial-number-input"),
    ]
    end = time.time() + timeout
    while time.time() < end:
        for by, val in strategies:
            try:
                el = driver.find_element(by, val)
                if el and el.is_displayed(): return el
            except: pass
        time.sleep(0.6)
    return None

def smooth_type(driver, el, text):
    try:
        el.click()
        el.send_keys(Keys.CONTROL + "a")
        el.send_keys(Keys.BACKSPACE)
        time.sleep(0.2)
        for char in text:
            el.send_keys(char)
            time.sleep(random.uniform(0.07, 0.15))
        time.sleep(0.3)
    except: pass

def find_model(driver):
    for xp in ["//h2", "//h1", "//*[contains(@class,'device') or contains(@class,'product')]"]:
        try:
            for el in driver.find_elements(By.XPATH, xp):
                t = el.text.strip()
                if t and any(k in t for k in ["iPhone", "iPad", "Watch", "MacBook", "iMac", "Mac", "AirPods", "iPod"]):
                    return t[:40]
        except: pass
    return ""

def find_date(line):
    m = re.search(r"(\d{1,2})\s+tháng\s+(\d{1,2}),\s*(\d{4})", line, re.I)
    if m: d, mt, y = m.groups(); return f"{int(d):02d}/{int(mt):02d}/{y}"
    return line.strip()

try:
    import ddddocr
    _ocr = ddddocr.DdddOcr(show_ad=False)            # model thuong
    try: _ocr2 = ddddocr.DdddOcr(show_ad=False, beta=True)   # model beta (doi chieu)
    except Exception: _ocr2 = None
    AUTO_CAPTCHA = True
except Exception:
    _ocr = _ocr2 = None; AUTO_CAPTCHA = False

def _find(driver, pairs):
    for by, val in pairs:
        try:
            el = driver.find_element(by, val)
            if el and el.is_displayed(): return el
        except: pass
    return None

def find_captcha_img(driver):
    return _find(driver, [
        (By.CSS_SELECTOR, ".captcha-container img"),
        (By.ID, "captcha-image"),
        (By.CSS_SELECTOR, "img.captcha-image, #captcha-image img"),
    ])

def find_captcha_input(driver):
    return _find(driver, [
        (By.CSS_SELECTOR, ".captcha-input-box input.form-textbox-input"),
        (By.CSS_SELECTOR, ".captcha-container .form-textbox-input"),
        (By.ID, "captcha-textbox"),
    ])

def find_continue(driver):
    # NUT GUI THAT = serial-button (type=submit, duy nhat 1 cai tren trang).
    # TUYET DOI KHONG dung captcha-action -> do la vung nut REFRESH (captcha-refresh-btn)!
    return _find(driver, [
        (By.CSS_SELECTOR, ".serial-button button"),
        (By.XPATH, "//div[contains(@class,'serial-button')]//button"),
        (By.CSS_SELECTOR, "button[type='submit']"),
        (By.XPATH, "//button[@type='submit' and (contains(.,'Gửi') or contains(.,'Continue') or contains(.,'Tiếp tục'))]"),
    ])

def find_refresh_captcha(driver):
    # Nut lam moi captcha THAT cua Apple
    return _find(driver, [
        (By.ID, "captcha-refresh-btn"),
        (By.CSS_SELECTOR, "#captcha-refresh-btn, .captcha-action button[type='button']"),
    ])

def _ocr_clean(txt):
    return re.sub(r'[^A-Za-z0-9]', '', (txt or '')).upper()

def read_captcha_text(driver, img):
    """OCR: anh goc base64 + grayscale + phong to + autocontrast (xu ly NHE, khong pha anh).
       2 model doi chieu. Tra (text, strong): strong=True khi 2 model cho CUNG ket qua."""
    data = None
    try:
        src = img.get_attribute("src") or ""
        if src.startswith("data:image"):
            import base64
            data = base64.b64decode(src.split(",", 1)[1])
    except: pass
    if not data:
        try: data = img.screenshot_as_png
        except: return "", False
    # Tien xu ly NHE (khong nhi phan/khu nhieu manh -> tranh pha net chu)
    proc = data
    try:
        from PIL import Image, ImageOps
        import io
        im = Image.open(io.BytesIO(data)).convert("L")
        w, h = im.size
        if w < 450: im = im.resize((w*3, h*3), Image.LANCZOS)
        im = ImageOps.autocontrast(im)
        buf = io.BytesIO(); im.save(buf, format="PNG"); proc = buf.getvalue()
    except: pass
    t1 = _ocr_clean(_ocr.classification(proc)) if _ocr else ""
    # CHI co 1 model -> tin neu dung dinh dang (khong cross-check duoc, tranh ket vong lap)
    if _ocr2 is None:
        return t1, (4 <= len(t1) <= 8)
    t2 = _ocr_clean(_ocr2.classification(proc))
    # STRONG = 2 model cho cung ket qua + dung do dai (rat it khi 2 model cung sai giong nhau)
    if t1 and t1 == t2 and 4 <= len(t1) <= 8:
        return t1, True
    cand = t1 if 4 <= len(t1) <= 8 else t2
    return cand, False

def captcha_sig(img):
    """Chu ky anh captcha -> de biet anh CO DOI hay khong (tranh doc lai anh cu)."""
    try:
        s = img.get_attribute("src") or ""
        if s: return s[-160:]
    except: pass
    try: return str(len(img.screenshot_as_png))
    except: return ""

def _captcha_loaded(img):
    """True neu anh captcha da LOAD XONG (base64 that, khong phai 'Loading')."""
    try:
        src = img.get_attribute("src") or ""
    except: return False
    if src.startswith("data:image"):
        return len(src) > 300        # data uri du dai = anh that
    # khong co src data -> dua vao kich thuoc
    try: return img.size.get('width', 0) > 20
    except: return False

def wait_captcha_ready(driver, old_sig=None, timeout=8):
    """ĐỢI captcha LOAD XONG (va KHAC anh cu neu old_sig). Tra (img, sig)."""
    end = time.time() + timeout
    while time.time() < end:
        if looks_banned(driver): return None, ""
        img = find_captcha_img(driver)
        if img and _captcha_loaded(img):
            sig = captcha_sig(img)
            if old_sig is None or sig != old_sig:
                return img, sig      # da load xong + la anh moi
        time.sleep(0.5)
    img = find_captcha_img(driver)
    return img, (captcha_sig(img) if img else "")

def refresh_and_wait(driver, old_sig, timeout=8):
    """BAM nut refresh roi ĐỢI anh moi LOAD XONG. Tra (img, sig)."""
    btn = find_refresh_captcha(driver)
    if btn:
        try: btn.click()
        except Exception:
            try: driver.execute_script("arguments[0].click()", btn)
            except: pass
    else:
        print("   [!] KHÔNG thấy nút làm mới captcha")
    img, sig = wait_captcha_ready(driver, old_sig=old_sig, timeout=timeout)
    if sig == old_sig:
        print("   [!] Ảnh captcha chưa đổi sau khi refresh")
    return img, sig

def looks_banned(driver):
    try: b = driver.find_element(By.TAG_NAME, "body").text.lower()
    except: return False
    return any(k in b for k in ["access denied", "forbidden", "reference #", "unusual activity", "too many requests"])

def submit_and_check(driver, cin_element, timeout=8):
    try: 
        cin_element.send_keys(Keys.ENTER)
        print("   [+] Đã gõ ENTER gửi mã...")
    except: pass
    
    time.sleep(1.0)
    
    btn = find_continue(driver)
    if btn:
        try: driver.execute_script("arguments[0].click()", btn)
        except:
            try: btn.click()
            except: pass
            
    end_time = time.time() + timeout
    while time.time() < end_time:
        try: 
            b = driver.find_element(By.TAG_NAME, "body").text.lower()
            if looks_banned(driver): return "BANNED"
            
            if any(k in b for k in ["không chính xác", "khong chinh xac", "thử lại", "sai mã", "incorrect", "try again"]):
                return "wrong_captcha"
                
            if any(k in b for k in ["số sê-ri không hợp lệ", "invalid serial", "không thể tìm thấy"]):
                return "success"
                
            s_box = find_serial_input(driver, timeout=1)
            if s_box is None or not s_box.is_displayed():
                if any(k in b for k in ["ngày mua", "kich hoat", "kích hoạt", "hợp lệ", "đã mua", "bảo hành"]):
                    return "success"
        except: pass
        time.sleep(0.5)
    return "CHECKING" 

# ==============================================================================
# BẢN VÁ v4.2: CHỐT CỨNG CHỈ GỬI KHI ĐỦ 4 KÝ TỰ
# ==============================================================================
def type_captcha(driver, cin, text):
    """Go ma vao o + KIEM TRA da vao chua, chua thi ep bang JS (React)."""
    smooth_type(driver, cin, text)
    time.sleep(0.4)
    try: cur = (driver.execute_script("return arguments[0].value;", cin) or "")
    except: cur = ""
    if cur.upper().replace(" ", "") != text:
        try:
            driver.execute_script(
                "var e=arguments[0],v=arguments[1];e.focus();e.value=v;"
                "['input','change','blur'].forEach(function(n){e.dispatchEvent(new Event(n,{bubbles:true}));});",
                cin, text)
            time.sleep(0.3)
        except: pass

APPLE_URL = "https://checkcoverage.apple.com/?locale=vi_VN"

def reload_and_enter_serial(driver, serial):
    """Tai LAI trang (cung IP) lay captcha MOI - KHONG bam refresh. Tra (img, sig)."""
    try: driver.get(APPLE_URL)
    except:
        try: driver.execute_script("window.stop();")
        except: pass
    time.sleep(2.5)
    if looks_banned(driver): return None, None
    el = find_serial_input(driver, timeout=12)
    if not el: return None, None
    smooth_type(driver, el, serial)
    time.sleep(1.5)
    return wait_captcha_ready(driver)

def solve_captcha_loop(driver, serial):
    """SU THAT: Apple BAN ngay khi gui DU 1 captcha SAI -> moi IP CHI duoc gui 1 lan.
       => Chua chac -> LOAD MA captcha moi (bam refresh, KHONG load lai trang, KHONG gui).
          Chỉ GỬI khi nhiều model + biến thể ĐỒNG THUẬN mạnh. Gửi sai -> ĐỔI IP ngay."""
    img, sig = wait_captcha_ready(driver)
    if img is None:
        if looks_banned(driver): return "BANNED"
        print("   [!] Không load được captcha"); return "ERROR"

    cin = find_captcha_input(driver)
    if cin:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", img)
            time.sleep(0.3)
        except: pass

    reads = 0
    while reads < CAPTCHA_TRIES:
        if looks_banned(driver): return "BANNED"
        cin = find_captcha_input(driver)
        if img is None or not cin:
            print("   [!] Mất ô/ảnh captcha -> đổi IP"); return False

        text, strong = read_captcha_text(driver, img)   # strong = 2 model khớp
        print(f"   captcha OCR (đọc {reads+1}/{CAPTCHA_TRIES}): {text} {'[CHẮC]' if strong else '[chưa chắc]'}")

        # CHUA CHAC -> chỉ LOAD MÃ MỚI (bấm refresh captcha), KHÔNG load lại trang, KHÔNG gửi
        if not (4 <= len(text) <= 8) or not strong:
            print("   [!] Chưa chắc -> load mã captcha mới (refresh, KHÔNG load lại trang)")
            img, sig = refresh_and_wait(driver, sig)
            reads += 1; continue

        # CHAC -> GUI DUY NHAT 1 LAN tren IP nay
        type_captcha(driver, cin, text)
        print(f"   -> Gửi mã: {text}")
        status = submit_and_check(driver, cin, timeout=10)
        if status == "success": return True
        if status == "BANNED": return "BANNED"
        print("   [!] Mã sai -> ĐỔI IP (IP này coi như đã cháy)")
        return False

    return False

def handle(driver, serial, i, total):
    try:
        print(f"\n({i}/{total}) Serial: {serial}")
        try:
            driver.get("https://checkcoverage.apple.com/?locale=vi_VN")
            time.sleep(3.0)
        except:
            try: driver.execute_script("window.stop();")
            except: pass

        if looks_banned(driver): return Result(serial, "BANNED", "", "", "")

        el = find_serial_input(driver, timeout=15)
        if not el:
            if looks_banned(driver): return Result(serial, "BANNED", "", "", "")
            return Result(serial, "ERROR", "", "", "")

        smooth_type(driver, el, serial)
        time.sleep(1.5)

        passed = solve_captcha_loop(driver, serial) if AUTO_CAPTCHA else None
        # Bat MOI ket qua loi cua captcha -> tra ve de main DOI IP + thu lai (khong ghi Unknown lung tung)
        if passed in ("TIMEOUT", "BANNED", "ERROR"): return Result(serial, passed, "", "", "")
        if passed is False:   # OCR sai het cac lan -> doi IP thu lai (captcha moi)
            print("   [!] Captcha chưa qua sau nhiều lần -> RETRY (đổi IP)")
            return Result(serial, "RETRY", "", "", "")

        time.sleep(1.5)
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        
        if any(k in body for k in ["số sê-ri không hợp lệ", "invalid serial", "không thể tìm thấy"]):
            print("Trạng thái: Invalid"); return Result(serial, "Invalid", "Không có", "", "")

        if any(k in body for k in ["ngày mua không hợp lệ", "chưa được kích hoạt", "chưa được xác nhận"]):
            print("Trạng thái: Unactivated"); return Result(serial, "Unactivated", find_model(driver), "", "")
            
        pur = exp = ""
        for line in driver.find_element(By.TAG_NAME, "body").text.split('\n'):
            if any(k in line.lower() for k in ["đã mua", "ngày mua"]): pur = find_date(line)
            if any(k in line.lower() for k in ["hết hạn", "dự kiến hết hạn"]): exp = find_date(line)
                
        mdl = find_model(driver)
        if pur or exp or mdl:
            print(f"Trạng thái: Activated | Model: {mdl} | Mua: {pur}")
            return Result(serial, "Activated", mdl, pur, exp)
            
        return Result(serial, "Unknown", "", "", "")
    except: return Result(serial, "ERROR", "", "", "")

def main():
    print("Apple Checker v5.4 - 1 IP/serial; chua chac thi LOAD MA captcha moi (refresh), KHONG load lai trang")
    master_serials = read_serials(INPUT_FILE)
    if not master_serials: return
    
    existing_results = load_existing_results(RESULT_FILE)
    serials_to_check = []
    for real_idx, s in enumerate(master_serials, start=1):
        if s not in existing_results or existing_results[s].status not in ("Activated", "Unactivated", "Invalid"):
            serials_to_check.append((real_idx, s))
                
    if not serials_to_check:
        print("\n===== TẤT CẢ MÃ ĐÃ HOÀN THÀNH SẠCH SẼ! ====="); return
    
    set_anti_sleep(True)
    retry_count = {}; MAX_RETRY = 4
    cur_proxy = fetch_rotating_proxy() if (PROXY_API_URL or PROXY_KEY) else (PROXY or None)
    
    if cur_proxy:
        print("   [+] Chờ 7 giây cho cổng proxy khởi động...")
        time.sleep(7)
        
    done_on_ip = 0
    try:
        idx = 0
        while idx < len(serials_to_check):
            real_idx, s = serials_to_check[idx]

            # Đổi IP định kỳ (sau mỗi BATCH_PER_IP serial)
            if done_on_ip >= BATCH_PER_IP:
                print(f"\n>>> Đã check {BATCH_PER_IP} mã trên IP này. Đổi IP...")
                if (PROXY_API_URL or PROXY_KEY):
                    new = fetch_rotating_proxy()
                    if new: cur_proxy = new
                print("   [+] Chờ 5 giây cho IP mới ổn định...")
                time.sleep(5)
                done_on_ip = 0

            # ===== MỖI SERIAL = 1 TRÌNH DUYỆT MỚI (trang sạch 100%, không dính state serial trước) =====
            d = create_browser(cur_proxy)
            try:
                r = handle(d, s, real_idx, len(master_serials))
            except Exception:
                r = Result(s, "ERROR", "", "", "")
            finally:
                try: d.quit()          # ĐÓNG trang ngay sau mỗi serial
                except: pass

            # Lỗi/ban -> đổi IP + thử lại CHÍNH serial này (vòng sau mở browser mới)
            if r.status in ("BANNED", "RETRY", "TIMEOUT", "ERROR"):
                retry_count[s] = retry_count.get(s, 0) + 1
                if retry_count[s] > MAX_RETRY:
                    existing_results[s] = r; idx += 1; done_on_ip += 1
                    save_results(master_serials, existing_results)
                    print(f"   [!] {s} thử {MAX_RETRY} lần không được -> bỏ qua ({r.status})")
                    continue
                print(f"   [!] Lỗi ({r.status}) -> đổi IP, thử lại sau 8s...")
                if (PROXY_API_URL or PROXY_KEY):
                    new = fetch_rotating_proxy()
                    if new: cur_proxy = new
                time.sleep(8)
                done_on_ip = 0
                continue   # KHÔNG tăng idx -> thử lại serial này

            # Thành công -> LƯU rồi serial tiếp theo
            existing_results[s] = r; idx += 1; done_on_ip += 1
            save_results(master_serials, existing_results)
            print(f"   -> Đã lưu kết quả ({r.status})")
            if idx < len(serials_to_check):
                time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))   # nghỉ giống người
    finally:
        set_anti_sleep(False)
        save_results(master_serials, existing_results)
        print("\n===== HOÀN THÀNH TASK =====")

if __name__ == "__main__": main()
